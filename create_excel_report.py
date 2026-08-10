import openpyxl
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

output = r"C:\Users\DELL LATITUDE 5520\.openclaw\workspace\temp_attachments\Sales_Report_Jan_Jul_2026.xlsx"
wb = openpyxl.Workbook()

# ---- DATA ----
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'July']
reps = ['Abdullah', 'Mustafa', 'Aamir', 'Siraj', 'Kashif', 'Asif']

data = {
    'Abdullah': [75000, 80000, 70000, 85000, 65000, 88000, 85000],
    'Mustafa':  [125000, 130000, 150000, 145000, 152000, 170000, 180000],
    'Aamir':    [25000, 35000, 60000, 45000, 50000, 52000, 58000],
    'Siraj':    [95000, 105000, 110000, 102000, 106000, 125000, 115000],
    'Kashif':   [35000, 33000, 32000, 36000, 33000, 32500, 38000],
    'Asif':     [46000, 39000, 47000, 48300, 50000, 49000, 53000],
}

monthly_totals = [401000, 422000, 469000, 461300, 456000, 516500, 529000]
rep_totals = {r: sum(data[r]) for r in reps}
total_all = sum(rep_totals.values())
avg_monthly = sum(monthly_totals) / len(monthly_totals)
ytd_target = 1200000

# ---- Styles ----
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(bold=True, size=14, color='2F5496')
subtitle_font = Font(bold=True, size=11, color='2F5496')
money_fmt = '#,##0'
pct_fmt = '0.0%'
center = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = center
    if fmt:
        cell.number_format = fmt

# ========================
# Sheet 1: Raw Data
# ========================
ws1 = wb.active
ws1.title = "Raw Data"

ws1.merge_cells('A1:J1')
ws1['A1'] = 'Sales Report — Jan to Jul 2026'
ws1['A1'].font = title_font

ws1.merge_cells('A2:J2')
ws1['A2'] = 'Source: Email from Syed Aamir Ali (aamir7601@gmail.com), Date: Aug 6, 2026'
ws1['A2'].font = Font(italic=True, color='666666')

# Headers
headers = ['Sales Rep'] + months + ['YTD Total', 'Monthly Avg', 'YTD Target']
for c, h in enumerate(headers, 1):
    ws1.cell(row=4, column=c, value=h)
style_header_row(ws1, 4, len(headers))

# Data rows
for i, rep in enumerate(reps):
    row = 5 + i
    ws1.cell(row=row, column=1, value=rep)
    ws1.cell(row=row, column=1).font = Font(bold=True)
    style_data_cell(ws1, row, 1)
    for j, val in enumerate(data[rep]):
        ws1.cell(row=row, column=2+j, value=val)
        style_data_cell(ws1, row, 2+j, money_fmt)
    total = rep_totals[rep]
    ws1.cell(row=row, column=9, value=total)
    style_data_cell(ws1, row, 9, money_fmt)
    ws1.cell(row=row, column=10, value=round(total/7))
    style_data_cell(ws1, row, 10, money_fmt)
    ws1.cell(row=row, column=11, value=ytd_target)
    style_data_cell(ws1, row, 11, money_fmt)

# Total row
total_row = 5 + len(reps)
ws1.cell(row=total_row, column=1, value='TOTAL')
ws1.cell(row=total_row, column=1).font = Font(bold=True, color='2F5496')
style_data_cell(ws1, total_row, 1)
for j, val in enumerate(monthly_totals):
    ws1.cell(row=total_row, column=2+j, value=val)
    style_data_cell(ws1, total_row, 2+j, money_fmt)
ws1.cell(row=total_row, column=9, value=total_all)
style_data_cell(ws1, total_row, 9, money_fmt)

# Column widths
ws1.column_dimensions['A'].width = 14
for c in range(2, 12):
    ws1.column_dimensions[get_column_letter(c)].width = 14

# ========================
# Sheet 2: Summary & Analysis
# ========================
ws2 = wb.create_sheet("Summary & Analysis")

ws2.merge_cells('A1:F1')
ws2['A1'] = 'Sales Report — Summary & Analysis'
ws2['A1'].font = title_font

# Key Metrics
ws2['A3'] = 'KEY METRICS'
ws2['A3'].font = subtitle_font

metrics = [
    ('Total Sales (Jan-Jul)', f'PKR {total_all:,}'),
    ('Monthly Average', f'PKR {avg_monthly:,.0f}'),
    ('Best Month', f'July — PKR {max(monthly_totals):,}'),
    ('Worst Month', f'January — PKR {min(monthly_totals):,}'),
    ('Growth (Jan → July)', f'{((monthly_totals[-1]-monthly_totals[0])/monthly_totals[0]*100):.1f}%'),
    ('YTD Target', f'PKR {ytd_target:,}'),
    ('Top Performer', f'Mustafa — PKR {rep_totals["Mustafa"]:,} (32.3%)'),
    ('Most Improved', f'Aamir — 132% growth (Jan to July)'),
]

for i, (label, value) in enumerate(metrics):
    ws2.cell(row=4+i, column=1, value=label).font = Font(bold=True)
    ws2.cell(row=4+i, column=3, value=value)
    ws2.merge_cells(f'C{4+i}:F{4+i}')

# Rep Ranking Table
rank_start = 14
ws2.cell(row=rank_start, column=1, value='REP PERFORMANCE RANKING')
ws2.cell(row=rank_start, column=1).font = subtitle_font

rank_headers = ['Rank', 'Sales Rep', 'YTD Sales', 'Share %', 'Monthly Avg', 'Trend']
for c, h in enumerate(rank_headers, 1):
    ws2.cell(row=rank_start+1, column=c, value=h)
style_header_row(ws2, rank_start+1, len(rank_headers))

sorted_reps = sorted(reps, key=lambda r: rep_totals[r], reverse=True)
for i, rep in enumerate(sorted_reps):
    row = rank_start + 2 + i
    total = rep_totals[rep]
    share = total / total_all
    ws2.cell(row=row, column=1, value=i+1)
    ws2.cell(row=row, column=2, value=rep).font = Font(bold=True)
    ws2.cell(row=row, column=3, value=total)
    ws2.cell(row=row, column=4, value=share)
    ws2.cell(row=row, column=5, value=round(total/7))
    # Trend
    first = data[rep][0]
    last = data[rep][-1]
    trend = (last - first) / first * 100
    trend_str = f'↑ {trend:.0f}%' if trend > 0 else f'↓ {abs(trend):.0f}%'
    ws2.cell(row=row, column=6, value=trend_str)
    for c in range(1, 7):
        style_data_cell(ws2, row, c)
    ws2.cell(row=row, column=3).number_format = money_fmt
    ws2.cell(row=row, column=4).number_format = pct_fmt
    ws2.cell(row=row, column=5).number_format = money_fmt
    # Highlight Mustafa
    if rep == 'Mustafa':
        for c in range(1, 7):
            ws2.cell(row=row, column=c).fill = green_fill

# Monthly breakdown
mon_start = rank_start + 2 + len(reps) + 2
ws2.cell(row=mon_start, column=1, value='MONTHLY BREAKDOWN')
ws2.cell(row=mon_start, column=1).font = subtitle_font

mon_headers = ['Month', 'Total Sales', 'vs Average', 'Growth MoM']
for c, h in enumerate(mon_headers, 1):
    ws2.cell(row=mon_start+1, column=c, value=h)
style_header_row(ws2, mon_start+1, len(mon_headers))

for i, (month, val) in enumerate(zip(months, monthly_totals)):
    row = mon_start + 2 + i
    ws2.cell(row=row, column=1, value=month).font = Font(bold=True)
    ws2.cell(row=row, column=2, value=val)
    diff = val - avg_monthly
    ws2.cell(row=row, column=3, value=f'{"+" if diff>=0 else ""}{diff/1000:.0f}K ({(diff/avg_monthly*100):.1f}%)')
    if i > 0:
        mom = (val - monthly_totals[i-1]) / monthly_totals[i-1] * 100
        ws2.cell(row=row, column=4, value=f'{"+" if mom>=0 else ""}{mom:.1f}%')
    for c in range(1, 5):
        style_data_cell(ws2, row, c)
    ws2.cell(row=row, column=2).number_format = money_fmt
    # Color based on vs average
    if diff >= 0:
        ws2.cell(row=row, column=3).fill = green_fill
    else:
        ws2.cell(row=row, column=3).fill = red_fill

# Column widths
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 28
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 14

# ========================
# Sheet 3: Charts
# ========================
ws3 = wb.create_sheet("Charts Data")

# Data for charts
ws3.cell(row=1, column=1, value='Month')
for j, m in enumerate(months):
    ws3.cell(row=1, column=2+j, value=m)
for i, rep in enumerate(reps):
    ws3.cell(row=2+i, column=1, value=rep)
    for j, val in enumerate(data[rep]):
        ws3.cell(row=2+i, column=2+j, value=val)

# Rep totals
ws3.cell(row=10, column=1, value='Rep')
ws3.cell(row=10, column=2, value='YTD Sales')
for i, rep in enumerate(reps):
    ws3.cell(row=11+i, column=1, value=rep)
    ws3.cell(row=11+i, column=2, value=rep_totals[rep])

# Monthly totals
ws3.cell(row=10, column=4, value='Month')
ws3.cell(row=10, column=5, value='Total Sales')
for i, (m, v) in enumerate(zip(months, monthly_totals)):
    ws3.cell(row=11+i, column=4, value=m)
    ws3.cell(row=11+i, column=5, value=v)

# ---- Chart 1: Monthly Trend Line ----
chart1 = LineChart()
chart1.title = "Monthly Sales Trend by Representative"
chart1.y_axis.title = "Sales (PKR)"
chart1.x_axis.title = "Month"
chart1.style = 10
chart1.height = 14
chart1.width = 22

cats = Reference(ws3, min_col=2, max_col=8, min_row=1, max_row=1)
for i in range(len(reps)):
    values = Reference(ws3, min_col=2, max_col=8, min_row=2+i, max_row=2+i)
    chart1.add_data(values, titles_from_data=False)
    chart1.series[i].title = openpyxl.chart.series.SeriesLabel(v=reps[i])
chart1.set_categories(cats)

ws3.add_chart(chart1, "A19")

# ---- Chart 2: YTD Bar Chart ----
chart2 = BarChart()
chart2.type = "col"
chart2.title = "YTD Sales by Representative"
chart2.y_axis.title = "Sales (PKR)"
chart2.style = 10
chart2.height = 14
chart2.width = 22

cats2 = Reference(ws3, min_col=1, max_col=1, min_row=11, max_row=16)
values2 = Reference(ws3, min_col=2, max_col=2, min_row=11, max_row=16)
chart2.add_data(values2, titles_from_data=False)
chart2.set_categories(cats2)
chart2.series[0].title = openpyxl.chart.series.SeriesLabel(v='YTD Sales')

ws3.add_chart(chart2, "A37")

# ---- Chart 3: Monthly Totals Bar ----
chart3 = BarChart()
chart3.type = "col"
chart3.title = "Total Monthly Sales (Jan-Jul 2026)"
chart3.y_axis.title = "Sales (PKR)"
chart3.style = 10
chart3.height = 14
chart3.width = 22

cats3 = Reference(ws3, min_col=4, max_col=4, min_row=11, max_row=17)
values3 = Reference(ws3, min_col=5, max_col=5, min_row=11, max_row=17)
chart3.add_data(values3, titles_from_data=False)
chart3.set_categories(cats3)
chart3.series[0].title = openpyxl.chart.series.SeriesLabel(v='Monthly Total')

ws3.add_chart(chart3, "A55")

# ---- Chart 4: Pie Chart ----
chart4 = PieChart()
chart4.title = "Sales Contribution by Representative"
chart4.height = 14
chart4.width = 14

cats4 = Reference(ws3, min_col=1, max_col=1, min_row=11, max_row=16)
values4 = Reference(ws3, min_col=2, max_col=2, min_row=11, max_row=16)
chart4.add_data(values4, titles_from_data=False)
chart4.set_categories(cats4)

chart4.dataLabels = DataLabelList()
chart4.dataLabels.showPercent = True
chart4.dataLabels.showCatName = True

ws3.add_chart(chart4, "A73")

# Save
wb.save(output)
print(f"Excel file saved: {output}")
print(f"Size: {os.path.getsize(output):,} bytes")
