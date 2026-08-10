import openpyxl
import json

filepath = r"C:\Users\DELL LATITUDE 5520\.openclaw\workspace\temp_attachments\test_sales_rep.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)

print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
        print(list(row))

wb.close()
